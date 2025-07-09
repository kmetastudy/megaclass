"""
Excel export utilities for Health Habit data
"""
from django.http import HttpResponse
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from app_home.models import DailyReflection
import re
from datetime import datetime
from urllib.parse import quote


class HealthHabitExcelExporter:
    """Excel exporter for Health Habit data"""
    
    def __init__(self, trackers, class_id=None):
        self.trackers = trackers
        self.class_id = class_id
        self.wb = None
        
        # 6가지 고정된 약속
        self.default_promises = {
            '1': '바른 자세로 생활하기',
            '2': '규칙적으로 가벼운 운동하기',
            '3': '바른 식습관 기르기',
            '4': '몸을 깨끗하게 하기',
            '5': '생활 주변을 깨끗하게 하기',
            '6': '마음 건강하게 관리하기'
        }
        
        # 요일 데이터
        self.days = ['월', '화', '수', '목', '금', '토', '일']
        
        # 스타일 정의
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 조건부 스타일
        self.submitted_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        self.not_submitted_fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
    
    def create_workbook(self):
        """Create Excel workbook with all sheets"""
        self.wb = Workbook()
        
        # 시트 생성
        self.create_student_overview_sheet()
        self.create_promise_details_sheet()
        self.create_daily_reflections_sheet()
        self.create_statistics_sheet()
        
        # 컬럼 너비 조정
        self.set_column_widths()
        
        return self.wb
    
    def create_student_overview_sheet(self):
        """학생 개요 시트 생성"""
        ws = self.wb.active
        ws.title = "학생 개요"
        
        # 헤더 작성
        headers = ['학번', '이름', '학년', '반', '제출여부', '완료율(%)', '총 소감수', '받은 별', '평가 등급']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 학생 데이터 작성
        for row, tracker in enumerate(self.trackers, 2):
            stats = tracker.get_completion_stats()
            
            # 받은 별 개수 계산
            total_stars = tracker.reflections.filter(
                evaluation__isnull=False
            ).aggregate(total=Sum('evaluation__score'))['total'] or 0
            
            # 학번에서 번호 추출
            student_number = tracker.student.student_id.split('_')[-1] if '_' in tracker.student.student_id else tracker.student.student_id
            
            # 평가 등급
            evaluation_grade = tracker.overall_evaluation.grade if hasattr(tracker, 'overall_evaluation') else ''
            
            data = [
                student_number,
                tracker.student.user.get_full_name(),
                tracker.student.school_class.grade,
                tracker.student.school_class.class_number,
                '제출' if tracker.is_submitted else '미제출',
                stats['completion_rate'],
                stats['total_reflections'],
                total_stars,
                evaluation_grade
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = self.center_alignment
                
                # 제출 여부에 따른 배경색
                if col == 5:  # 제출여부 컬럼
                    if tracker.is_submitted:
                        cell.fill = self.submitted_fill
                    else:
                        cell.fill = self.not_submitted_fill
    
    def create_promise_details_sheet(self):
        """약속 세부사항 시트 생성"""
        ws = self.wb.create_sheet("약속 세부사항")
        
        # 헤더 작성
        headers = ['학번', '이름']
        for i in range(1, 7):
            headers.extend([f'약속{i}', f'약속{i} 실천방법'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 약속 데이터 작성
        for row, tracker in enumerate(self.trackers, 2):
            student_number = tracker.student.student_id.split('_')[-1] if '_' in tracker.student.student_id else tracker.student.student_id
            
            data = [student_number, tracker.student.user.get_full_name()]
            
            for i in range(1, 7):
                promise_title = tracker.promises.get(str(i), self.default_promises.get(str(i), f'약속 {i}'))
                promise_detail = tracker.promise_details.get(str(i), '')
                data.extend([promise_title, promise_detail])
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 2:  # 약속 내용은 왼쪽 정렬
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment
        
        # 약속 제목과 실천방법 개별 헤더 생성
        for i in range(1, 7):
            promise_col = 3 + (i-1) * 2
            detail_col = promise_col + 1
            
            # 약속 제목 컬럼 (예: "약속1")
            promise_cell = ws.cell(row=1, column=promise_col, value=f'약속{i}')
            promise_cell.font = self.header_font
            promise_cell.fill = self.header_fill
            promise_cell.alignment = self.center_alignment
            promise_cell.border = self.border
            
            # 세부사항 컬럼 (예: "세부사항")
            detail_cell = ws.cell(row=1, column=detail_col, value='세부사항')
            detail_cell.font = self.header_font
            detail_cell.fill = self.header_fill
            detail_cell.alignment = self.center_alignment
            detail_cell.border = self.border
    
    def create_daily_reflections_sheet(self):
        """일일 소감 시트 생성 (완전한 데이터 포함)"""
        ws = self.wb.create_sheet("일일 소감")
        
        # 헤더 작성
        headers = ['학번', '이름', '약속번호', '약속제목', '주차', '요일', '날짜', '소감내용', '평가점수', '교사피드백']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 완전한 데이터 생성
        reflection_data = self.generate_complete_reflection_data()
        
        # 데이터 작성
        current_row = 2
        prev_student = None
        student_start_row = 2
        
        for data_row in reflection_data:
            student_number = data_row['student_number']
            student_name = data_row['student_name']
            
            # 학생별 셀 병합을 위한 시작 행 추적
            if prev_student != student_name:
                if prev_student is not None:
                    # 이전 학생의 셀 병합
                    if current_row - 1 > student_start_row:
                        ws.merge_cells(f'A{student_start_row}:A{current_row-1}')
                        ws.merge_cells(f'B{student_start_row}:B{current_row-1}')
                student_start_row = current_row
                prev_student = student_name
            
            data = [
                student_number,
                student_name,
                data_row['promise_number'],
                data_row['promise_title'],
                data_row['week'],
                data_row['day_name'],
                data_row['date'],
                data_row['reflection_text'],
                data_row['evaluation_score'],
                data_row['evaluation_comment']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self.border
                
                if col in [8, 10]:  # 소감내용, 교사피드백은 왼쪽 정렬
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment
                
                # 제출된 소감과 미제출 소감 구분
                if data_row['is_submitted']:
                    cell.fill = self.submitted_fill
                else:
                    cell.fill = self.not_submitted_fill
            
            current_row += 1
        
        # 마지막 학생의 셀 병합
        if prev_student is not None and current_row - 1 > student_start_row:
            ws.merge_cells(f'A{student_start_row}:A{current_row-1}')
            ws.merge_cells(f'B{student_start_row}:B{current_row-1}')
    
    def create_statistics_sheet(self):
        """통계 시트 생성"""
        ws = self.wb.create_sheet("통계")
        
        # 약속별 통계
        ws.cell(row=1, column=1, value="약속별 완료율 통계").font = Font(bold=True, size=14)
        
        stat_headers = ['약속번호', '약속제목', '총 소감수', '완료율(%)']
        for col, header in enumerate(stat_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        for promise_num in range(1, 7):
            total_reflections = DailyReflection.objects.filter(
                tracker__in=self.trackers,
                promise_number=promise_num
            ).count()
            
            total_possible = self.trackers.count() * 14
            completion_rate = (total_reflections / total_possible * 100) if total_possible > 0 else 0
            
            promise_title = self.default_promises.get(str(promise_num), f'약속 {promise_num}')
            
            data = [promise_num, promise_title, total_reflections, round(completion_rate, 1)]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=promise_num + 3, column=col, value=value)
                cell.border = self.border
                cell.alignment = self.center_alignment
    
    def generate_complete_reflection_data(self):
        """모든 가능한 소감 데이터 생성 (기본값 포함)"""
        complete_data = []
        
        for tracker in self.trackers:
            student_number = tracker.student.student_id.split('_')[-1] if '_' in tracker.student.student_id else tracker.student.student_id
            student_name = tracker.student.user.get_full_name()
            
            # 기존 소감 데이터를 딕셔너리로 변환 (빠른 조회를 위해)
            existing_reflections = {}
            for reflection in tracker.reflections.all():
                key = f"{reflection.promise_number}_{reflection.week}_{reflection.day}"
                existing_reflections[key] = reflection
            
            # 모든 가능한 조합 생성 (6 약속 × 2 주차 × 7 요일 = 84개)
            for promise_num in range(1, 7):
                promise_title = tracker.promises.get(str(promise_num), 
                                                   self.default_promises.get(str(promise_num), 
                                                                           f'약속 {promise_num}'))
                
                for week in range(1, 3):  # 1주차, 2주차
                    for day in range(1, 8):  # 1-7 (월-일)
                        key = f"{promise_num}_{week}_{day}"
                        day_name = self.days[day - 1]
                        
                        # 기존 데이터가 있는지 확인
                        if key in existing_reflections:
                            reflection = existing_reflections[key]
                            
                            # 평가 정보
                            evaluation_score = ''
                            evaluation_comment = ''
                            if hasattr(reflection, 'evaluation'):
                                evaluation_score = reflection.evaluation.score
                                evaluation_comment = reflection.evaluation.comment
                            
                            complete_data.append({
                                'student_number': student_number,
                                'student_name': student_name,
                                'promise_number': promise_num,
                                'promise_title': promise_title,
                                'week': week,
                                'day_name': day_name,
                                'date': reflection.reflection_date.strftime('%Y-%m-%d') if reflection.reflection_date else '',
                                'reflection_text': reflection.reflection_text,
                                'evaluation_score': evaluation_score,
                                'evaluation_comment': evaluation_comment,
                                'is_submitted': True
                            })
                        else:
                            # 기본값으로 빈 데이터 생성
                            complete_data.append({
                                'student_number': student_number,
                                'student_name': student_name,
                                'promise_number': promise_num,
                                'promise_title': promise_title,
                                'week': week,
                                'day_name': day_name,
                                'date': '',
                                'reflection_text': '미작성',
                                'evaluation_score': '-',
                                'evaluation_comment': '-',
                                'is_submitted': False
                            })
        
        # 요일 순서 매핑 (월요일부터 일요일까지)
        weekday_order = {'월': 1, '화': 2, '수': 3, '목': 4, '금': 5, '토': 6, '일': 7}
        
        # 정렬: 학생번호 → 약속번호 → 주차 → 요일 순
        complete_data.sort(key=lambda x: (
            x['student_number'],  # 학생번호로 정렬 (학생명 대신)
            x['promise_number'], 
            x['week'], 
            weekday_order.get(x['day_name'], 8)  # 요일을 올바른 순서로 정렬
        ))
        
        return complete_data
    
    def calculate_text_width(self, text):
        """텍스트 너비 계산 (한글 고려)"""
        if not text:
            return 10
        
        # 한글 문자 수 계산
        korean_chars = len(re.findall(r'[가-힣]', str(text)))
        # 영문/숫자/기호 문자 수 계산
        other_chars = len(str(text)) - korean_chars
        
        # 한글은 2배 너비로 계산
        width = korean_chars * 2 + other_chars
        
        # 최소 15, 최대 50으로 제한
        return min(max(width, 15), 50)
    
    def set_column_widths(self):
        """컬럼 너비 동적 조정"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            # 각 컬럼별로 최대 너비 계산
            column_widths = {}
            
            for row in ws.iter_rows():
                for cell in row:
                    # MergedCell 체크 - MergedCell은 건너뛰기
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.column_letter not in column_widths:
                        column_widths[cell.column_letter] = 0
                    
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            cell_width = self.calculate_text_width(cell.value)
                            if cell_width > column_widths[cell.column_letter]:
                                column_widths[cell.column_letter] = cell_width
                    except:
                        pass
            
            # 컬럼 너비 적용
            for column_letter, max_length in column_widths.items():
                # 최소 너비 15, 최대 너비 50으로 조정
                adjusted_width = min(max(max_length, 15), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_http_response(self):
        """HTTP 응답 생성"""
        wb = self.create_workbook()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 파일명 생성 (날짜 정보 포함)
        current_date = datetime.now().strftime('%Y%m%d')
        
        if self.class_id:
            if self.trackers.exists():
                school_class = self.trackers.first().student.school_class
                class_name = f'{school_class.grade}-{school_class.class_number}반'
            else:
                class_name = '전체'
            filename = f'건강습관_데이터_{class_name}_{current_date}.xlsx'
        else:
            filename = f'건강습관_데이터_전체_{current_date}.xlsx'
        
        # 한글 파일명을 위한 UTF-8 인코딩 (RFC 6266 방식)
        filename_encoded = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename_encoded}'
        
        wb.save(response)
        return response