"""
PAPS Excel Export 유틸리티 함수들
openpyxl을 사용하여 Excel 파일을 생성하는 범용 함수 제공
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from django.http import HttpResponse
from .models import PAPSActivity
from .utils import DemoSessionManager, get_grade_display
import io


def create_paps_excel_workbook(headers, rows, filename="paps_export.xlsx"):
    """
    PAPS Excel 워크북 생성 함수 (데이터 소스 무관)
    
    Args:
        headers (List[str]): 컬럼 헤더 리스트 ['학년도', '학년', '반명', ...]
        rows (List[List[Any]]): 데이터 행들 [[2024, '초4', '1반', ...], ...]
        filename (str): 다운로드할 파일명
        
    Returns:
        HttpResponse: Excel 파일 다운로드 응답
    """
    # 새 워크북 생성
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "PAPS측정데이터"
    
    # 스타일 정의
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더 행 추가
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = center_alignment
    
    # 데이터 행 추가
    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.border = border_thin
            cell.alignment = center_alignment
    
    # DimensionHolder 방식으로 동적 컬럼 너비 자동 조정
    dim_holder = DimensionHolder(worksheet=worksheet)
    
    # 각 컬럼별 최대 너비 계산
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # 해당 컬럼의 모든 셀 순회
        for row in range(worksheet.min_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=col)
            try:
                if cell.value is not None:
                    # 한글 텍스트 너비 계산 함수 활용
                    cell_length = _calculate_korean_text_width(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass
        
        # 적절한 패딩과 최소/최대 너비 적용
        # 최소 8, 최대 50, 패딩 +2
        adjusted_width = min(max(max_length + 2, 8), 50)
        
        # ColumnDimension 객체 생성 및 할당
        dim_holder[column_letter] = ColumnDimension(worksheet, min=col, max=col, width=adjusted_width)
    
    # 워크시트에 dimension 적용
    worksheet.column_dimensions = dim_holder
    
    # Excel 파일을 메모리에 저장
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    # HTTP 응답 생성
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def prepare_demo_data_for_excel(request):
    """
    체험판 데이터를 Excel 형식으로 준비
    
    Args:
        request: Django request 객체 (세션 데이터 접근용)
        
    Returns:
        tuple: (headers, rows) - Excel 생성에 필요한 헤더와 데이터
    """
    # 가상 학생 데이터 가져오기
    demo_students = DemoSessionManager.get_demo_students()
    
    # 세션에서 모든 측정 기록 가져오기
    session_records = DemoSessionManager.get_session_records(request)
    
    # 측정된 종목들 식별
    measured_activities = set()
    for record in session_records.values():
        measured_activities.add(record['activity_name'])
    
    # 측정된 종목들의 PAPSActivity 정보 조회
    activities = PAPSActivity.objects.filter(name__in=measured_activities).order_by('name')
    
    # 헤더 구성 (성별 제거, 반코드 추가)
    headers = ['학년도', '학년', '반명', '반코드', '번호', '학생성명']
    
    # 측정 항목별 필드 헤더 추가
    measurement_fields = []
    for activity in activities:
        if activity.measurement_schema and 'fields' in activity.measurement_schema:
            for field in activity.measurement_schema['fields']:
                if not field.get('readonly', False):  # 입력 가능한 필드만
                    field_info = {
                        'activity_name': activity.name,
                        'field_key': field['field'],
                        'field_name': field['name']
                    }
                    measurement_fields.append(field_info)
                    headers.append(field['name'])
    
    # 데이터 행 구성
    rows = []
    for student in demo_students:
        row = [
            2024,  # 학년도 (고정)
            get_grade_display(student['grade']),  # 학년 (get_grade_display 함수 사용)
            '체험반',  # 반명 (고정)
            '',  # 반코드 (빈 값)
            student['number'],
            student['name']
        ]
        
        # 각 측정 필드별 데이터 추가
        for field_info in measurement_fields:
            # 해당 학생의 해당 종목 기록 찾기
            record_key = DemoSessionManager._get_record_key(student['id'], field_info['activity_name'])
            
            if record_key in session_records:
                measurement_data = session_records[record_key]['measurement_data']
                field_value = measurement_data.get(field_info['field_key'], '')
            else:
                field_value = ''
            
            row.append(field_value)
        
        rows.append(row)
    
    return headers, rows


def _calculate_korean_text_width(text):
    """
    한글 텍스트의 Excel 컬럼 너비를 계산
    한글/한자는 1.5배, 영문/숫자는 1배로 계산
    """
    if not text:
        return 0
    
    width = 0
    for char in str(text):
        # 한글, 한자, 일본어 등 동아시아 문자
        if '\u3000' <= char <= '\u303f' or \
           '\u3040' <= char <= '\u309f' or \
           '\u30a0' <= char <= '\u30ff' or \
           '\uff00' <= char <= '\uff9f' or \
           '\u4e00' <= char <= '\u9faf' or \
           '\u3400' <= char <= '\u4dbf' or \
           '\uac00' <= char <= '\ud7a3':
            width += 1.5
        else:
            width += 1
    
    return int(width)